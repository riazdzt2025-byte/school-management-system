from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
from .models import (
    Student, Subject, Institution, TransferCertificate, Certificate,
    SSCRegistration, BoardResult, Exam, ExamMark,
)
from .forms import (
    StudentForm, SubjectForm, ExcelImportForm, TransferCertificateForm,
    CertificateForm, SSCRegistrationForm, BoardResultForm, SSCExcelImportForm,
    ExamForm,
)
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse
import openpyxl
from collections import defaultdict
from .result_utils import build_exam_results



@login_required
def dashboard(request):
    total_students = Student.objects.count()
    total_subjects = Subject.objects.count()
    return render(request, 'students/dashboard.html', {
        'total_students': total_students,
        'total_subjects': total_subjects,
    })
@login_required
def class_section_summary(request):
    institutions = Institution.objects.all().order_by('name')
    institution_id = request.GET.get('institution')
    institution = None

    students_qs = Student.objects.all()
    if institution_id:
        institution = get_object_or_404(Institution, pk=institution_id)
        students_qs = students_qs.filter(institution=institution)

    summary = defaultdict(lambda: {'total': 0, 'male': 0, 'female': 0, 'other': 0})

    for s in students_qs.only('admission_class', 'section', 'gender'):
        key = (s.admission_class, s.section)
        summary[key]['total'] += 1
        if s.gender == 'M':
            summary[key]['male'] += 1
        elif s.gender == 'F':
            summary[key]['female'] += 1
        else:
            summary[key]['other'] += 1

    summary_rows = []
    for (cls, section), counts in summary.items():
        summary_rows.append({
            'admission_class': cls,
            'section': section,
            **counts,
        })

    summary_rows.sort(key=lambda r: (r['admission_class'], r['section']))

    grand_total = students_qs.count()

    return render(request, 'students/class_section_summary.html', {
        'institutions': institutions,
        'institution': institution,
        'summary_rows': summary_rows,
        'grand_total': grand_total,
    })
#---------------- Student Views ----------------
@login_required
def student_list(request):
    institutions = Institution.objects.all().order_by('name')
    institutions_data = {
        str(inst.id): [c.strip() for c in inst.classes.split(',') if c.strip()]
        for inst in institutions
    }

    institution = None
    admission_class = request.GET.get('admission_class')
    section = request.GET.get('section')
    institution_id = request.GET.get('institution')
    show_filter_modal = False

    if request.GET.get('all') == '1':
        students = list(Student.objects.all())
    elif institution_id:
        institution = get_object_or_404(Institution, pk=institution_id)
        qs = Student.objects.filter(institution=institution)
        if admission_class:
            qs = qs.filter(admission_class=admission_class)
        if section:
            qs = qs.filter(section__iexact=section.strip())
        students = list(qs)
    else:
        students = []
        show_filter_modal = True

    # ---- Duplicate detection ----
    exact_key_count = defaultdict(int)
    name_count = defaultdict(int)
    for s in students:
        name_key = s.name.strip().lower()
        exact_key_count[(name_key, s.admission_class, s.section)] += 1
        name_count[name_key] += 1

    exact_duplicate_ids = set()
    possible_duplicate_ids = set()
    for s in students:
        name_key = s.name.strip().lower()
        if exact_key_count[(name_key, s.admission_class, s.section)] > 1:
            exact_duplicate_ids.add(s.id)
        elif name_count[name_key] > 1:
            possible_duplicate_ids.add(s.id)

    return render(request, 'students/student_list.html', {
        'students': students,
        'institution': institution,
        'selected_class': admission_class,
        'selected_section': section,
        'institutions': institutions,
        'institutions_data': institutions_data,
        'show_filter_modal': show_filter_modal,
        'exact_duplicate_ids': exact_duplicate_ids,
        'possible_duplicate_ids': possible_duplicate_ids,
    })


@login_required
@permission_required('students.delete_student', raise_exception=True)
def bulk_delete_students(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        deleted_count = Student.objects.filter(pk__in=student_ids).delete()[0]
        messages.success(request, f"{deleted_count} student(s) deleted.")

        url = reverse('student_list')
        params = []
        if request.POST.get('institution'):
            params.append(f"institution={request.POST.get('institution')}")
        if request.POST.get('admission_class'):
            params.append(f"admission_class={request.POST.get('admission_class')}")
        if request.POST.get('section'):
            params.append(f"section={request.POST.get('section')}")
        if params:
            url += "?" + "&".join(params)
        return redirect(url)
    return redirect('student_list')
@login_required
@permission_required('students.change_student', raise_exception=True)
def bulk_update_students(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        new_class = request.POST.get('new_class', '').strip()
        new_section = request.POST.get('new_section', '').strip()

        if not student_ids:
            messages.error(request, "No students were selected.")
        elif not new_class and not new_section:
            messages.error(request, "Please provide a new Class or Section to update.")
        else:
            qs = Student.objects.filter(pk__in=student_ids)
            update_fields = {}
            if new_class:
                update_fields['admission_class'] = new_class
            if new_section:
                update_fields['section'] = new_section
            updated_count = qs.update(**update_fields)
            messages.success(request, f"{updated_count} student(s) updated successfully.")

        url = reverse('student_list')
        params = []
        if request.POST.get('institution'):
            params.append(f"institution={request.POST.get('institution')}")
        if request.POST.get('admission_class'):
            params.append(f"admission_class={request.POST.get('admission_class')}")
        if request.POST.get('section'):
            params.append(f"section={request.POST.get('section')}")
        if params:
            url += "?" + "&".join(params)
        return redirect(url)
    return redirect('student_list')

@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'students/student_detail.html', {
        'student': student,
        'subjects': student.subjects.select_related('subject'),
        'transfer_certificate': getattr(student, 'transfer_certificate', None),
    })


@login_required
@permission_required('students.add_transfercertificate', raise_exception=True)
def issue_tc(request, pk):
    student = get_object_or_404(Student, pk=pk)
    existing_tc = getattr(student, 'transfer_certificate', None)
    if existing_tc:
        messages.info(request, "This student's TC has already been issued.")
        return redirect('view_tc', pk=existing_tc.pk)

    if request.method == 'POST':
        form = TransferCertificateForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                transfer_certificate = form.save(commit=False)
                transfer_certificate.student = student
                transfer_certificate.save()
                student.status = 'TRANSFERRED'
                student.save(update_fields=['status'])
            messages.success(request, f"Transfer Certificate {transfer_certificate.tc_number} issued.")
            return redirect('view_tc', pk=transfer_certificate.pk)
    else:
        form = TransferCertificateForm()

    return render(request, 'students/issue_tc.html', {
        'student': student,
        'form': form,
    })


@login_required
def view_tc(request, pk):
    transfer_certificate = get_object_or_404(TransferCertificate, pk=pk)
    return render(request, 'students/tc_print.html', {
        'tc': transfer_certificate,
        'student': transfer_certificate.student,
        'transfer_certificate': transfer_certificate,
    })


@login_required
@permission_required('students.add_certificate', raise_exception=True)
def issue_certificate(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = CertificateForm(request.POST)
        if form.is_valid():
            certificate = form.save(commit=False)
            certificate.student = student
            certificate.save()
            messages.success(request, f"Certificate issued — {certificate.certificate_number}")
            return redirect('view_certificate', pk=certificate.pk)
    else:
        form = CertificateForm()

    return render(request, 'students/issue_certificate.html', {
        'form': form,
        'student': student,
    })


@login_required
def view_certificate(request, pk):
    certificate = get_object_or_404(Certificate, pk=pk)
    return render(request, 'students/certificate_print.html', {
        'cert': certificate,
        'student': certificate.student,
    })


@login_required
def certificate_list(request, pk):
    student = get_object_or_404(Student, pk=pk)
    certificates = student.certificates.all().order_by('-issue_date', '-pk')
    return render(request, 'students/certificate_list.html', {
        'student': student,
        'certificates': certificates,
    })


@login_required
@permission_required('students.add_student', raise_exception=True)
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            try:
                student = form.save()
                messages.success(request, f"Student added — ID: {student.student_id}")
                if 'save_add_another' in request.POST:
                    url = reverse('add_student')
                    if student.institution_id:
                        url += f"?institution={student.institution_id}"
                    return redirect(url)
                url = reverse('student_list')
                if student.institution_id:
                    url += f"?institution={student.institution_id}"
                return redirect(url)
            except ValueError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "There are errors in the form — please check the fields below.")
    else:
        form = StudentForm()
        institution_id = request.GET.get('institution')
        if institution_id:
            form.fields['institution'].initial = institution_id
    return render(request, 'students/add_student.html', {'form': form})


@login_required
@permission_required('students.change_student', raise_exception=True)
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student information updated.")
            url = reverse('student_list')
            if student.institution_id:
                url += f"?institution={student.institution_id}"
            return redirect(url)
        else:
            messages.error(request, "There are errors in the form — please check the fields below.")
    else:
        form = StudentForm(instance=student)
    return render(request, 'students/add_student.html', {'form': form, 'student': student})


@login_required
@permission_required('students.delete_student', raise_exception=True)
def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, "Student deleted.")
        return redirect('student_list')
    return render(request, 'students/delete_student.html', {'student': student})


# ---------------- Subject Views ----------------

@login_required
def subject_list(request):
    subjects = Subject.objects.all()
    return render(request, 'students/subject_list.html', {
        'subjects': subjects,
    })


@login_required
@permission_required('students.add_subject', raise_exception=True)
def add_subject(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject added.")
            return redirect('subject_list')
        else:
            messages.error(request, "There are errors in the form — please check the fields below.")
    else:
        form = SubjectForm()
    return render(request, 'students/add_subject.html', {'form': form})


@login_required
@permission_required('students.change_subject', raise_exception=True)
def edit_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject information updated.")
            return redirect('subject_list')
        else:
            messages.error(request, "There are errors in the form — please check the fields below.")
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'students/add_subject.html', {'form': form, 'subject': subject})


@login_required
@permission_required('students.delete_subject', raise_exception=True)
def delete_subject(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, "Subject deleted.")
        return redirect('subject_list')
    return render(request, 'students/delete_subject.html', {'subject': subject})


# ---------------- Excel Import ----------------

@login_required
@permission_required('students.add_student', raise_exception=True)
def import_students(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
            except Exception as e:
                messages.error(request, f"Could not read the file: {e}")
                return render(request, 'students/import_students.html', {'form': form})

            success_count = 0
            error_rows = []

            gender_map = {'male': 'M', 'm': 'M',
                          'female': 'F', 'f': 'F',
                          'other': 'O', 'o': 'O'}
            group_map = {label.lower(): code for code, label in Student.GROUP_CHOICES}

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell in (None, '') for cell in row):
                    continue
                try:
                    row_data = (list(row) + [None] * 12)[:12]
                    (institution_name, name, admission_class, section, admission_year,
                     roll_no, gender_raw, religion, father_name, contact_no,
                     guardian_contact_no, group_raw) = row_data

                    if not name or not admission_class:
                        error_rows.append(f"Row {row_num}: name or class is empty — skipped.")
                        continue

                    institution = None
                    if institution_name:
                        institution = Institution.objects.filter(
                            name__iexact=str(institution_name).strip()
                        ).first()
                        if not institution:
                            error_rows.append(
                                f"Row {row_num}: institution '{institution_name}' not found — skipped."
                            )
                            continue

                    gender_code = gender_map.get(str(gender_raw).strip().lower(), '') if gender_raw else ''
                    group_code = group_map.get(str(group_raw).strip().lower(), '') if group_raw else ''

                    Student.objects.create(
                        institution=institution,
                        name=str(name).strip(),
                        admission_class=str(admission_class).strip(),
                        section=str(section).strip() if section else '',
                        admission_year=int(admission_year) if admission_year else 0,
                        roll_no=int(roll_no) if roll_no else 0,
                        gender=gender_code,
                        religion=str(religion).strip() if religion else '',
                        father_name=str(father_name).strip() if father_name else '',
                        contact_no=str(contact_no).strip() if contact_no else '',
                        guardian_contact_no=str(guardian_contact_no).strip() if guardian_contact_no else '',
                        group=group_code,
                    )
                    success_count += 1
                except Exception as e:
                    error_rows.append(f"Row {row_num}: error — {e}")

            if success_count:
                messages.success(request, f"{success_count} student(s) added successfully.")
            if error_rows:
                messages.warning(
                    request,
                    f"{len(error_rows)} row(s) had issues: " + " | ".join(error_rows[:10])
                )
            return redirect('student_list')
    else:
        form = ExcelImportForm()
    return render(request, 'students/import_students.html', {'form': form})


@login_required
@permission_required('students.add_student', raise_exception=True)
def download_import_template(request):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Students"

    headers = [
        "Institution", "Name", "Admission Class", "Section",
        "Admission Year", "Roll No", "Gender", "Religion",
        "Father's Name", "Contact No", "Guardian Contact No", "Group",
    ]
    sheet.append(headers)

    sheet.append([
        "Principal Kazi Faruky School", "Example Name", "6", "A",
        2026, 1, "Male", "Islam", "Father's Name", "01700000000",
        "01800000000", "Non-Group",
    ])

    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response


# ---------------- SSC Registration Views ----------------

@login_required
def ssc_registration_list(request):
    registrations = SSCRegistration.objects.select_related('student', 'board_result').all()
    session = request.GET.get('session')
    if session:
        registrations = registrations.filter(session=session)
    sessions = SSCRegistration.objects.values_list('session', flat=True).distinct().order_by('session')
    return render(request, 'students/ssc_registration_list.html', {
        'registrations': registrations,
        'sessions': sessions,
        'selected_session': session,
    })


@login_required
@permission_required('students.add_sscregistration', raise_exception=True)
def add_ssc_registration(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if hasattr(student, 'ssc_registration'):
        messages.info(request, "This student is already registered for SSC.")
        return redirect('ssc_registration_list')
    form = SSCRegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        registration = form.save(commit=False)
        registration.student = student
        registration.save()
        messages.success(request, f"SSC registration added — {registration.registration_number}")
        return redirect('ssc_registration_list')
    return render(request, 'students/add_ssc_registration.html', {'form': form, 'student': student})


@login_required
@permission_required('students.change_sscregistration', raise_exception=True)
def edit_ssc_registration(request, pk):
    registration = get_object_or_404(SSCRegistration, pk=pk)
    form = SSCRegistrationForm(request.POST or None, instance=registration)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "SSC registration updated.")
        return redirect('ssc_registration_list')
    return render(request, 'students/add_ssc_registration.html', {
        'form': form, 'student': registration.student, 'registration': registration,
    })


@login_required
@permission_required('students.delete_sscregistration', raise_exception=True)
def delete_ssc_registration(request, pk):
    registration = get_object_or_404(SSCRegistration, pk=pk)
    if request.method == 'POST':
        registration.delete()
        messages.success(request, "SSC registration deleted.")
        return redirect('ssc_registration_list')
    return render(request, 'students/delete_ssc_registration.html', {'registration': registration})


@login_required
@permission_required('students.add_sscregistration', raise_exception=True)
def import_ssc_registrations(request):
    form = SSCExcelImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        try:
            sheet = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True).active
        except Exception as exc:
            messages.error(request, f"Could not read the file: {exc}")
            return render(request, 'students/import_ssc_registrations.html', {'form': form})

        group_map = {label.lower(): code for code, label in SSCRegistration.GROUP_CHOICES}
        board_map = {label.lower(): code for code, label in SSCRegistration.BOARD_CHOICES}
        success_count, error_rows = 0, []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell in (None, '') for cell in row):
                continue
            try:
                values = (list(row) + [None] * 7)[:7]
                student_id, reg_no, roll_no, session, group_raw, subjects, board_raw = values
                student = Student.objects.filter(student_id=str(student_id).strip()).first() if student_id else None
                group_code = group_map.get(str(group_raw).strip().lower()) if group_raw else None
                board_code = board_map.get(str(board_raw).strip().lower()) if board_raw else None
                if not student or not reg_no or not group_code or not board_code:
                    raise ValueError('student ID, registration number, group, or board is invalid')
                if hasattr(student, 'ssc_registration'):
                    raise ValueError('student is already registered')
                SSCRegistration.objects.create(
                    student=student, registration_number=str(reg_no).strip(),
                    roll_number=str(roll_no).strip() if roll_no else '',
                    session=str(session).strip() if session else '', group=group_code,
                    subjects=str(subjects).strip() if subjects else '', board=board_code,
                )
                success_count += 1
            except Exception as exc:
                error_rows.append(f"Row {row_num}: {exc}")
        if success_count:
            messages.success(request, f"{success_count} SSC registration(s) added successfully.")
        if error_rows:
            messages.warning(request, f"{len(error_rows)} row(s) had issues: " + ' | '.join(error_rows[:10]))
        return redirect('ssc_registration_list')
    return render(request, 'students/import_ssc_registrations.html', {'form': form})


# ---------------- Board Result Views ----------------

@login_required
@permission_required('students.add_boardresult', raise_exception=True)
def add_board_result(request, pk):
    registration = get_object_or_404(SSCRegistration, pk=pk)
    if hasattr(registration, 'board_result'):
        messages.info(request, "Result for this student has already been published.")
        return redirect('result_summary')
    form = BoardResultForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        result = form.save(commit=False)
        result.ssc_registration = registration
        result.save()
        messages.success(request, f"Result published for {registration.student.name}")
        return redirect('result_summary')
    return render(request, 'students/add_board_result.html', {'form': form, 'registration': registration})


@login_required
@permission_required('students.change_boardresult', raise_exception=True)
def edit_board_result(request, pk):
    result = get_object_or_404(BoardResult, pk=pk)
    form = BoardResultForm(request.POST or None, instance=result)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Result updated.")
        return redirect('result_summary')
    return render(request, 'students/add_board_result.html', {
        'form': form, 'registration': result.ssc_registration, 'result': result,
    })


@login_required
def ssc_result_summary(request):
    session = request.GET.get('session')
    registrations = SSCRegistration.objects.select_related('board_result', 'student').all()
    if session:
        registrations = registrations.filter(session=session)
    sessions = SSCRegistration.objects.values_list('session', flat=True).distinct().order_by('session')
    results = [registration.board_result for registration in registrations if hasattr(registration, 'board_result')]
    pass_count = sum(result.result_status == 'PASS' for result in results)
    gpa_buckets = {'5.00': 0, '4.00-4.99': 0, '3.50-3.99': 0, '3.00-3.49': 0, 'Below 3.00': 0}
    for result in results:
        if result.gpa is None:
            continue
        gpa = float(result.gpa)
        key = '5.00' if gpa >= 5 else '4.00-4.99' if gpa >= 4 else '3.50-3.99' if gpa >= 3.5 else '3.00-3.49' if gpa >= 3 else 'Below 3.00'
        gpa_buckets[key] += 1
    return render(request, 'students/result_summary.html', {
        'sessions': sessions, 'selected_session': session,
        'total_registered': registrations.count(), 'total_published': len(results),
        'pass_count': pass_count, 'fail_count': len(results) - pass_count,
        'pass_rate': round(pass_count / len(results) * 100, 2) if results else 0,
        'gpa_buckets': gpa_buckets,
        'a_plus_students': [result for result in results if result.grade == 'A+'],
    })


# ---------------- Exam Views ----------------

@login_required
def exam_list(request):
    exams = Exam.objects.all().order_by('-session', 'admission_class', 'name')
    return render(request, 'students/exam_list.html', {'exams': exams})


@login_required
@permission_required('students.add_exam', raise_exception=True)
def add_exam(request):
    form = ExamForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Exam created.')
        return redirect('exam_list')
    return render(request, 'students/add_exam.html', {'form': form})


@login_required
@permission_required('students.change_exam', raise_exception=True)
def edit_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    form = ExamForm(request.POST or None, instance=exam)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Exam updated.')
        return redirect('exam_list')
    return render(request, 'students/add_exam.html', {'form': form, 'exam': exam})


@login_required
@permission_required('students.delete_exam', raise_exception=True)
def delete_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method == 'POST':
        exam.delete()
        messages.success(request, 'Exam deleted.')
        return redirect('exam_list')
    return render(request, 'students/delete_exam.html', {'exam': exam})


@login_required
@permission_required('students.change_exam', raise_exception=True)
def toggle_publish_exam(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    exam.is_published = not exam.is_published
    exam.save(update_fields=['is_published'])
    status = 'published' if exam.is_published else 'unpublished'
    messages.success(request, f'Exam result {status}.')
    return redirect('exam_list')


@login_required
@permission_required('students.add_exammark', raise_exception=True)
def select_marks_subject(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    subjects = Subject.objects.all().order_by('name')
    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        if subject_id:
            return redirect('enter_marks', pk=exam.pk, subject_pk=subject_id)
        messages.error(request, 'Please select a subject.')
    return render(request, 'students/select_marks_subject.html', {'exam': exam, 'subjects': subjects})


@login_required
@permission_required('students.add_exammark', raise_exception=True)
def enter_marks(request, pk, subject_pk):
    exam = get_object_or_404(Exam, pk=pk)
    subject = get_object_or_404(Subject, pk=subject_pk)
    students_qs = Student.objects.filter(admission_class=exam.admission_class)
    if exam.section:
        students_qs = students_qs.filter(section__iexact=exam.section.strip())
    students = students_qs.order_by('roll_no', 'name')
    existing_marks = dict(ExamMark.objects.filter(exam=exam, subject=subject).values_list('student_id', 'marks_obtained'))

    if request.method == 'POST':
        saved_count = 0
        skipped_count = 0
        for student in students:
            value = request.POST.get(f'marks_{student.pk}', '').strip()
            if value == '':
                continue
            try:
                marks_value = float(value)
                if marks_value < 0 or marks_value > subject.full_marks:
                    skipped_count += 1
                    continue
            except ValueError:
                skipped_count += 1
                continue
            ExamMark.objects.update_or_create(
                exam=exam, student=student, subject=subject,
                defaults={'marks_obtained': marks_value},
            )
            saved_count += 1
        messages.success(request, f'Marks saved for {saved_count} student(s).')
        if skipped_count:
            messages.warning(request, f'{skipped_count} invalid mark(s) were skipped.')
        return redirect('exam_list')

    students_with_marks = [(student, existing_marks.get(student.pk, '')) for student in students]
    return render(request, 'students/enter_marks.html', {
        'exam': exam,
        'subject': subject,
        'students_with_marks': students_with_marks,
    })


# ---------------- Exam Result Views ----------------

@login_required
def result_sheet(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    subjects, results = build_exam_results(exam)
    return render(request, 'students/result_sheet.html', {'exam': exam, 'subjects': subjects, 'results': results})


@login_required
def result_summary(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    _, results = build_exam_results(exam)
    return render(request, 'students/exam_result_summary.html', {'exam': exam, 'results': results})


@login_required
def top_10(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    _, results = build_exam_results(exam)
    return render(request, 'students/top10.html', {'exam': exam, 'results': [r for r in results if r['position']][:10]})


def _exam_result(exam, student_pk):
    student = get_object_or_404(Student, pk=student_pk)
    _, results = build_exam_results(exam)
    return student, next((r for r in results if r['student'].pk == student.pk), None)


@login_required
def student_result_detail(request, pk, student_pk):
    exam = get_object_or_404(Exam, pk=pk)
    student, result = _exam_result(exam, student_pk)
    if not result or not result['has_marks']:
        messages.error(request, 'No marks found for this student in this exam.')
        return redirect('exam_result_summary', pk=exam.pk)
    return render(request, 'students/student_result_detail.html', {'exam': exam, 'result': result})


@login_required
def result_card(request, pk, student_pk):
    exam = get_object_or_404(Exam, pk=pk)
    student, result = _exam_result(exam, student_pk)
    if not result or not result['has_marks']:
        messages.error(request, 'No marks found for this student in this exam.')
        return redirect('exam_result_summary', pk=exam.pk)
    return render(request, 'students/result_card.html', {'exam': exam, 'result': result})